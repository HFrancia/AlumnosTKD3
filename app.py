from flask import Flask, render_template, request, jsonify, send_file
from sqlalchemy import create_engine, Column, Integer, String, Date, ForeignKey, Float, JSON
from sqlalchemy.ext.declarative import declarative_base
from sqlalchemy.orm import sessionmaker, relationship
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Image as PDFImage
from io import BytesIO

app = Flask(__name__)

# Database setup
engine = create_engine('sqlite:///AlumnosTB.db', echo=True)
Base = declarative_base()
Session = sessionmaker(bind=engine)

class Alumno(Base):
    __tablename__ = 'alumnos'
    id = Column(Integer, primary_key=True)
    apaterno = Column(String(50), nullable=False)
    apmaterno = Column(String(50), nullable=False)
    nombre = Column(String(50), nullable=False)
    fbday = Column(Date, nullable=False)
    curp = Column(String(18), unique=True, nullable=False)
    calle = Column(String(100), nullable=True)
    numero = Column(String(10), nullable=False)
    colonia = Column(String(100), nullable=False)
    email = Column(String(100), nullable=False)
    telefono = Column(String(15), nullable=False)
    numafiliacion = Column(String(20), unique=True, nullable=True)
    estatus = Column(String(10), nullable=False)
    pagos = relationship("Pago", back_populates="alumno")

class Pago(Base):
    __tablename__ = 'pagos'
    id = Column(Integer, primary_key=True)
    alumno_id = Column(Integer, ForeignKey('alumnos.id'))
    fecha = Column(Date, nullable=False)
    monto = Column(Float, nullable=False)
    concepto = Column(String(100), nullable=False)
    alumno = relationship("Alumno", back_populates="pagos")

class Pedido(Base):
    __tablename__ = 'pedidos'
    id = Column(Integer, primary_key=True)
    fecha = Column(Date, nullable=False)
    nombre_solicitante = Column(String(100), nullable=False)
    tipo_producto = Column(String(50), nullable=False)
    tallas = Column(JSON, nullable=False)
    color = Column(String(50))
    cantidad = Column(Integer, nullable=False)
    costo_total = Column(Float, nullable=False)

Base.metadata.create_all(engine)

@app.route('/')
def index():
    return render_template('index.html')


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        session = Session()
        try:
            nuevo_alumno = Alumno(
                apaterno=request.form['apaterno'],
                apmaterno=request.form['apmaterno'],
                nombre=request.form['nombre'],
                fbday=datetime.strptime(request.form['fbday'], '%Y-%m-%d').date(),
                curp=request.form['curp'],
                calle=request.form['calle'],
                numero=request.form['numero'],
                colonia=request.form['colonia'],
                email=request.form['email'],
                telefono=request.form['telefono'],
                numafiliacion=request.form['numafiliacion'],
                estatus=request.form['estatus']
            )
            session.add(nuevo_alumno)
            session.commit()
            return jsonify({"success": True, "message": "Alumno registrado correctamente"})
        except Exception as e:
            session.rollback()
            return jsonify({"success": False, "message": f"Error: {str(e)}"})
        finally:
            session.close()
    return render_template('registro.html')

@app.route('/registro_equipo', methods=['GET', 'POST'])
def registro_equipo():
    if request.method == 'POST':
        session = Session()
        try:
            nuevo_pedido = Pedido(
                fecha=datetime.now().date(),
                nombre_solicitante=request.form['nombre_solicitante'],
                tipo_producto=request.form['tipo_producto'],
                tallas=request.form.getlist('tallas'),
                color=request.form.get('color'),
                cantidad=sum([int(request.form.get(f'cantidad_{talla}', 0)) for talla in request.form.getlist('tallas')]),
                costo_total=float(request.form['costo_total'])
            )
            session.add(nuevo_pedido)
            session.commit()
            return jsonify({"success": True, "message": "Pedido registrado correctamente"})
        except Exception as e:
            session.rollback()
            return jsonify({"success": False, "message": f"Error: {str(e)}"})
        finally:
            session.close()
    return render_template('registro_equipo.html')

@app.route('/pedidos', methods=['GET'])
def pedidos():
    session = Session()
    try:
        fecha_busqueda = request.args.get('fecha')
        if fecha_busqueda:
            fecha_busqueda = datetime.strptime(fecha_busqueda, '%Y-%m-%d').date()
            pedidos = session.query(Pedido).filter(Pedido.fecha == fecha_busqueda).all()
        else:
            pedidos = session.query(Pedido).all()
        return render_template('pedidos.html', pedidos=pedidos)
    except Exception as e:
        return f"Error: {str(e)}"
    finally:
        session.close()

@app.route('/alumnos')
def lista_alumnos():
    session = Session()
    try:
        alumnos = session.query(Alumno).all()
        return render_template('lista_alumnos.html', alumnos=alumnos)
    except Exception as e:
        return f"Error: {str(e)}"
    finally:
        session.close()

@app.route('/alumno/<int:id>', methods=['GET', 'POST'])
def detalle_alumno(id):
    session = Session()
    try:
        alumno = session.query(Alumno).get(id)
        if request.method == 'POST':
            alumno.apaterno = request.form['apaterno']
            alumno.apmaterno = request.form['apmaterno']
            alumno.nombre = request.form['nombre']
            alumno.fbday = datetime.strptime(request.form['fbday'], '%Y-%m-%d').date()
            alumno.curp = request.form['curp']
            alumno.calle = request.form['calle']
            alumno.numero = request.form['numero']
            alumno.colonia = request.form['colonia']
            alumno.email = request.form['email']
            alumno.telefono = request.form['telefono']
            alumno.numafiliacion = request.form['numafiliacion']
            alumno.estatus = request.form['estatus']
            session.commit()
            return jsonify({"success": True, "message": "Alumno actualizado correctamente"})
        return render_template('detalle_alumno.html', alumno=alumno)
    except Exception as e:
        session.rollback()
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    finally:
        session.close()

@app.route('/eliminar_alumno/<int:id>', methods=['POST'])
def eliminar_alumno(id):
    session = Session()
    try:
        alumno = session.query(Alumno).get(id)
        session.delete(alumno)
        session.commit()
        return jsonify({"success": True, "message": "Alumno eliminado correctamente"})
    except Exception as e:
        session.rollback()
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    finally:
        session.close()

@app.route('/pago/<int:alumno_id>', methods=['GET', 'POST'])
def pago(alumno_id): #redefinir variable a setpagos para identificar que es el metodo para aplicar pagos
    session = Session()
    try:
        alumno = session.query(Alumno).get(alumno_id)
        if request.method == 'POST':
            nuevo_pago = Pago(
                alumno_id=alumno_id,
                fecha=datetime.strptime(request.form['fecha'], '%Y-%m-%d').date(),
                monto=float(request.form['monto']),
                concepto=request.form['concepto']
            )
            session.add(nuevo_pago)
            session.commit()
            return jsonify({"success": True, "message": "Pago registrado correctamente"})
        return render_template('pago.html', alumno=alumno)
    except Exception as e:
        session.rollback()
        return jsonify({"success": False, "message": f"Error: {str(e)}"})
    finally:
        session.close()

@app.route('/pagos/<int:alumno_id>')
def pagos(alumno_id): #redefinit variable a getpagos para identificar que es el listado de pagos
    session = Session()
    try:
        alumno = session.query(Alumno).get(alumno_id)
        pagos = session.query(Pago).filter_by(alumno_id=alumno_id).all()
        return render_template('pagos.html', alumno=alumno, pagos=pagos)
    except Exception as e:
        return f"Error: {str(e)}"
    finally:
        session.close()
    
@app.route('/generar_reporte')
def generar_reporte():
    session = Session()
    try:
        alumnos = session.query(Alumno).filter(Alumno.estatus == "activo").all()
        df = pd.DataFrame([
            {
               'No': a.id,
               'Apellido Paterno': a.apaterno,
               'Apellido Materno': a.apmaterno,
               'Nombre': a.nombre,
               'Fecha de Nacimiento': a.fbday,
               'CURP': a.curp,
               'Calle': a.calle,
               'Número': a.numero,
               'Colonia': a.colonia,
               'Email': a.email,
               'Teléfono': a.telefono,
               'Número de Afiliación': a.numafiliacion
            } for a in alumnos
        ])

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Alumnos"

        # Add logo
        img = Image('static/img/logo_excl.png')
        img.width = 470  # Adjust as needed
        img.height = 80  # Adjust as needed
        ws.add_image(img, 'I1')

        # Merge cells for the logo
        ws.merge_cells('I1:L4')

        # Add generation date
        #ws['E1'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d')}"

        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for r, row in enumerate(df.values, start=6):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        filename = f"Reporte_Alumnos_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        return send_file(filename, as_attachment=True)
    except Exception as e:
        return f"Error: {str(e)}"
    finally:
        session.close()

@app.route('/generar_reporte_pagos/<int:alumno_id>')
def generar_reporte_pagos(alumno_id):
    session = Session()
    try:
        alumno = session.query(Alumno).get(alumno_id)
        pagos = session.query(Pago).filter_by(alumno_id=alumno_id).all()
        
        df = pd.DataFrame([
            {
                'Fecha Pago': p.fecha,
                'Monto': p.monto,
                'Concepto': p.concepto
            } for p in pagos
        ])

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Pagos"

        # Add logo
        img = Image('static/img/logo.png')
        img.width = 270  # Adjust as needed
        img.height = 80  # Adjust as needed
        ws.add_image(img, 'A1')

        # Merge cells for the logo
        ws.merge_cells('A1:A3')

        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for r, row in enumerate(df.values, start=6):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        filename = f"Reporte_Pagos_{alumno.nombre}_{alumno.apaterno}_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        return send_file(filename, as_attachment=True)
    except Exception as e:
        return f"Error: {str(e)}"
    finally:
        session.close()

@app.route('/generar_reporte_pedidos_excel')
def generar_reporte_pedidos_excel():
    session = Session()
    try:
        pedidos = session.query(Pedido).all()
        
        df = pd.DataFrame([
            {
                'Fecha': p.fecha,
                'Solicitante': p.nombre_solicitante,
                'Producto': p.tipo_producto,
                'Tallas': ', '.join(p.tallas),
                'Color': p.color or 'N/A',
                'Cantidad': p.cantidad,
                'Costo Total': p.costo_total
            } for p in pedidos
        ])

        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte de Pedidos"

        # Add logo
        img = Image('static/img/logo.png')
        img.width = 300
        img.height = 100
        ws.add_image(img, 'A1')

        # Merge cells for the logo
        ws.merge_cells('A1:D4')

        # Add generation date
        ws['E1'] = f"Fecha de generación: {datetime.now().strftime('%Y-%m-%d')}"

        # Write headers
        headers = list(df.columns)
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=5, column=col, value=header)
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = PatternFill(start_color="000080", end_color="000080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for r, row in enumerate(df.values, start=6):
            for c, value in enumerate(row, start=1):
                ws.cell(row=r, column=c, value=value)

        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        filename = f"Reporte_Pedidos_{datetime.now().strftime('%Y%m%d')}.xlsx"
        wb.save(filename)

        return send_file(filename, as_attachment=True)
    except Exception as e:
        return f"Error: {str(e)}"
    finally:
        session.close()

#Generación de reportes pdf
@app.route('/generar_reporte_pedidos_pdf')
def generar_reporte_pedidos_pdf():
    session = Session()
    try:
        pedidos = session.query(Pedido).all()
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []

        # Add logo
        logo = PDFImage('static/img/logo.png', width=100, height=50)
        elements.append(logo)

        # Create table data
        data = [['Producto', 'Tallas', 'Color', 'Cantidad', 'Costo Total']]
        for pedido in pedidos:
            data.append([
                pedido.tipo_producto,
                ', '.join(pedido.tallas),
                pedido.color or 'N/A',
                pedido.cantidad,
                f"${pedido.costo_total:.2f}"
            ])

        # Create table
        table = Table(data)
        style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.blue),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 14),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 12),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ])
        table.setStyle(style)
        elements.append(table)

        # Build PDF
        doc.build(elements)

        buffer.seek(0)
        return send_file(buffer, as_attachment=True, download_name='reporte_pedidos.pdf', mimetype='application/pdf')
    except Exception as e:
        return f"Error: {str(e)}"
    finally:
        session.close()

if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0')